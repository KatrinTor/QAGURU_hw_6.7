import os.path
from openpyxl import load_workbook
from conftest import RES_DIR


# TODO оформить в тест, добавить ассерты и использовать универсальный путь
def test_read_xlsx_file():
    xlsx_file = os.path.join(RES_DIR, 'file_example_XLSX_50.xlsx')
    workbook = load_workbook(xlsx_file)
    sheet = workbook.active
    headers = [cell.value for cell in sheet[1]]

    assert headers == [0, 'First Name', 'Last Name', 'Gender', 'Country', 'Age', 'Date', 'Id']
    assert sheet.cell(row=3, column=2).value == 'Mara'